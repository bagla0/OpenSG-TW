"""
Export the upper-surface OML stress comparison (MSG-TW vs FEniCS solid) to xlsx.

For each matched OML node (upper surface, TE -> LE) the sheet lists the TW and
the matched FEniCS-solid (y2,y3) coordinates, the match distance, and the six
3D stress components from each method with the percent error.

Reads outputs/oml_jax.txt and outputs/oml_fenics_gauss.txt (produced by
benchmark_oml_jax.py and the WSL FEniCS run).  Writes
outputs/oml_stress_comparison.xlsx.
"""
import os
import sys
import numpy as np
from scipy.spatial import cKDTree
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

sys.path.insert(0, os.path.dirname(__file__))
from benchmark_oml_compare import load, solid_oml_coords, upper_te_to_le, COMP, SOLID

OUT = os.path.join(os.path.dirname(__file__), "..", "outputs")
FF = "[1e5, 5e4, 5e4, 5e4, 1e5, 1e5]"   # the common beam force used (VABS)


def match_outer_idx(jxy, fxy, oml, radius=0.04):
    """Index (into fxy) of the outermost near solid point for each TW node."""
    cen = oml.mean(axis=0)
    rad_f = np.hypot(fxy[:, 0] - cen[0], fxy[:, 1] - cen[1])
    tree = cKDTree(fxy)
    midx = np.zeros(len(jxy), int); dist = np.zeros(len(jxy))
    for i, p in enumerate(jxy):
        cand = tree.query_ball_point(p, radius)
        j = cand[int(np.argmax(rad_f[cand]))] if cand else tree.query(p)[1]
        midx[i] = j; dist[i] = float(np.hypot(*(fxy[j] - p)))
    return midx, dist


# --------------------------------------------------------------------- data
ap_tw = os.path.join(OUT, "oml_tw_atpath.txt")
ap_fe = os.path.join(OUT, "oml_fenics_atpath.txt")
if os.path.exists(ap_tw) and os.path.exists(ap_fe):
    # EXACT-PATH mode: TW and solid evaluated at the SAME coordinates (dist~0)
    tw = np.loadtxt(ap_tw); fe = np.loadtxt(ap_fe)
    txy, js_all = tw[:, :2], tw[:, 2:]
    fexy, fes = fe[:, :2], fe[:, 2:]
    ii = cKDTree(txy).query(fexy)[1]               # align solid rows to TW path rows
    jxy, js = txy[ii], js_all[ii]
    dist = np.hypot(jxy[:, 0] - fexy[:, 0], jxy[:, 1] - fexy[:, 1])
    MODE = "EXACT path — TW and solid evaluated at the SAME coordinates"
else:
    # MATCH mode: TW nodes matched to the nearest near-OML solid Gauss point
    jxy, js = load(os.path.join(OUT, "oml_jax.txt"))
    idx = upper_te_to_le(jxy)
    jxy, js = jxy[idx], js[idx]                    # upper surface, TE -> LE
    oml = solid_oml_coords(SOLID)
    fxy, fs = load(os.path.join(OUT, "oml_fenics_gauss.txt"))
    near = cKDTree(oml).query(fxy)[0] < 0.05
    fxy, fs = fxy[near], fs[near]
    midx, dist = match_outer_idx(jxy, fxy, oml)
    fexy, fes = fxy[midx], fs[midx]
    MODE = "MATCH — TW nodes matched to nearest near-OML solid Gauss point"
print(MODE, f"({len(jxy)} nodes, max match dist {dist.max():.4f})")
s11_scale = float(np.max(np.abs(js[:, 0])))
peak = np.maximum(np.max(np.abs(fes), axis=0), 1e-30)   # per-component peak |solid|


# --------------------------------------------------------------------- xlsx
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "OML stress TW vs solid"

hdr = Font(bold=True, color="FFFFFF")
fill = PatternFill("solid", fgColor="305496")
sub = PatternFill("solid", fgColor="8EA9DB")
thin = Border(*([Side(style="thin", color="BFBFBF")] * 4))
center = Alignment(horizontal="center")

# title rows
ws["A1"] = f"MSG-TW (JAX) vs FEniCS solid (2Dsolid_0) — upper surface, TE->LE  [{MODE}]"
ws["A1"].font = Font(bold=True, size=12)
ws["A2"] = (f"common beam force FF = {FF} (VABS);  global frame;  "
            "%err = 100*(TW-solid)/solid;  err%pk = 100*(TW-solid)/max|solid_comp|")
ws["A3"] = ("%err blows up at zero-crossings (use err%pk there);  S22/S33/S23 ~1% "
            f"of S11 (free-surface ~0 in TW, sub-surface in solid);  S11={s11_scale:.3e} Pa")

# group header (row 5) + column header (row 6)
base = ["idx", "TW_y2", "TW_y3", "FE_y2", "FE_y3", "dist"]
r_grp, r_col = 5, 6
for c, name in enumerate(base, start=1):
    ws.cell(r_col, c, name).font = hdr
    ws.cell(r_col, c).fill = fill
    ws.cell(r_col, c).alignment = center
col = len(base) + 1
for comp in COMP:
    ws.cell(r_grp, col, comp).font = Font(bold=True)
    ws.cell(r_grp, col).alignment = center
    ws.merge_cells(start_row=r_grp, start_column=col, end_row=r_grp, end_column=col + 3)
    for k, sh in enumerate(["TW (Pa)", "solid (Pa)", "%err", "err%pk"]):
        cell = ws.cell(r_col, col + k, sh)
        cell.font = hdr; cell.fill = sub; cell.alignment = center
    col += 4

# data rows
r0 = r_col + 1
for i in range(len(jxy)):
    row = [i, jxy[i, 0], jxy[i, 1], fexy[i, 0], fexy[i, 1], dist[i]]
    for c, v in enumerate(row, start=1):
        cell = ws.cell(r0 + i, c, float(v) if c > 1 else int(v))
        cell.border = thin
        if c in (2, 3, 4, 5, 6):
            cell.number_format = "0.0000"
    col = len(base) + 1
    for j in range(6):
        tw, fe = float(js[i, j]), float(fes[i, j])
        err = (tw - fe) / fe * 100.0 if abs(fe) > 1e-6 * s11_scale else float("nan")
        errpk = (tw - fe) / peak[j] * 100.0
        for k, v in enumerate((tw, fe, err, errpk)):
            cell = ws.cell(r0 + i, col + k, v)
            cell.border = thin
            cell.number_format = "0.000E+00" if k < 2 else "0.0"
        col += 4

# summary: robust per-component median |err%pk| + median |err| as % of S11
rs = r0 + len(jxy) + 1
ws.cell(rs + 1, 1, "median |err%pk|:").font = Font(bold=True, italic=True)
ws.cell(rs + 2, 1, "median |err|(% S11):").font = Font(bold=True, italic=True)
col = len(base) + 1
for j, comp in enumerate(COMP):
    ws.cell(rs, col, comp).font = Font(bold=True)
    med_pk = float(np.median(np.abs((js[:, j] - fes[:, j]) / peak[j]) * 100))
    med_s11 = float(np.median(np.abs((js[:, j] - fes[:, j]) / s11_scale) * 100))
    ws.cell(rs + 1, col, med_pk).number_format = "0.00"
    ws.cell(rs + 2, col, med_s11).number_format = "0.00"
    col += 4

# cosmetics
ws.freeze_panes = "B7"
for c in range(1, len(base) + 25):
    ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 12

path = os.path.join(OUT, "oml_stress_comparison.xlsx")
wb.save(path)
print(f"wrote {path}  ({len(jxy)} matched OML nodes, 6 stress components)")
# quick console echo of the dominant components
print("\n idx  TW_y2  TW_y3 | S11 %err  S13 %err  S12 %err")
for i in range(len(jxy)):
    e = [ (js[i,k]-fes[i,k])/fes[i,k]*100 if abs(fes[i,k])>1e-6*s11_scale else float('nan')
          for k in (0,4,5) ]
    print(f"  {i:2d} {jxy[i,0]:6.2f} {jxy[i,1]:6.2f} | "
          f"{e[0]:7.1f}  {e[1]:7.1f}  {e[2]:7.1f}")
